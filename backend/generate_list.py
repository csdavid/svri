import sys
import openpyxl

sys.path.insert(0, '.')
from app.services.excel_parser import _classify_column

path = '/Users/david/trabajo-anam/proyectos/validacion-factura-impresion/documentacion/01-abril-2026.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

max_c = ws.max_column
row1 = [ws.cell(1, c).value for c in range(1, max_c + 1)]
row2 = [ws.cell(2, c).value for c in range(1, max_c + 1)]

for rng in ws.merged_cells.ranges:
    v1 = ws.cell(rng.min_row, rng.min_col).value
    for r in range(rng.min_row, min(rng.max_row + 1, 3)):
        for c in range(rng.min_col, rng.max_col + 1):
            if r == 1: row1[c-1] = v1
            elif r == 2: row2[c-1] = v1

headers = [ws.cell(2, c).value for c in range(1, max_c + 1)]

from collections import defaultdict
sections_map = defaultdict(list)

for c in range(max_c):
    r1_val = str(row1[c]).strip() if row1[c] else ''
    r2_val = str(row2[c]).strip() if row2[c] else ''
    h_val = str(headers[c]).strip() if headers[c] else f'COL_{c+1}'
    sec, sub = _classify_column(r1_val, r2_val, h_val, c + 1, False)
    sections_map[sec].append((c + 1, h_val, sub))

section_order = [
    'EQUIPO',
    'IMPRESION CARTA BN',
    'IMPRESION OFICIO BN',
    'IMPRESION DOBLECARTA BN',
    'DIGITALIZACION',
    'IMPRESION CARTA COLOR',
    'IMPRESION OFICIO COLOR',
    'IMPRESION DOBLECARTA COLOR',
    'LECTURAS INICIALES',
    'LECTURAS FINALES',
    'VOLUMEN',
    'PRECIOS',
    'IMPORTES FACTURACION',
    'TOTALES'
]

print(f"Total Columnas Detectadas: {max_c}\n")

for idx, sec in enumerate(section_order, 1):
    cols = sections_map.get(sec, [])
    print(f"### {idx}. SECCIÓN: {sec} ({len(cols)} columnas)")
    if not cols:
        print("*(Sin columnas asignadas)*\n")
        continue
    for col_idx, h, sub in cols:
        h_clean = h.replace('\n', ' ')
        print(f"* **Col {col_idx}**: `{h_clean}`")
    print()
