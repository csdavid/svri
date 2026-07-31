import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime
from sqlalchemy.orm import Session
from app.db.database import get_db
from app.models.models import LecturaFacturacionImpresion, LecturaFacturacionImpresionColumna

router = APIRouter(prefix="/api", tags=["Validación Manual"])

class ColumnaEstadoDTO(BaseModel):
    id_columna: int
    es_valido: bool

class ValidacionManualRequest(BaseModel):
    id_registro: int
    fecha_valida: bool
    decision: str  # 'ACEPTAR' or 'RECHAZAR'
    usuario: str = "admin"
    columnas_estados: List[ColumnaEstadoDTO]

@router.get("/periodos")
def get_periodos(db: Session = Depends(get_db)):
    results = db.query(
        LecturaFacturacionImpresion.periodo_anoi,
        LecturaFacturacionImpresion.periodo_mes
    ).distinct().all()
    
    return [
        {"periodo_anoi": r[0], "periodo_mes": r[1]} for r in results
    ]

@router.get("/series")
def get_series(
    periodo_anoi: str = Query(...),
    periodo_mes: str = Query(...),
    db: Session = Depends(get_db)
):
    series = db.query(LecturaFacturacionImpresion.serie).filter(
        LecturaFacturacionImpresion.periodo_anoi == periodo_anoi,
        LecturaFacturacionImpresion.periodo_mes == periodo_mes
    ).distinct().all()
    
    return [s[0] for s in series]

@router.get("/periodo-estatus")
def get_periodo_estatus(
    periodo_anoi: str = Query(...),
    periodo_mes: str = Query(...),
    db: Session = Depends(get_db)
):
    query = db.query(LecturaFacturacionImpresion).filter(
        LecturaFacturacionImpresion.periodo_anoi == periodo_anoi,
        LecturaFacturacionImpresion.periodo_mes == periodo_mes
    )

    total = query.count()
    validados = query.filter(LecturaFacturacionImpresion.validacion_manual.isnot(None)).count()
    pendientes = query.filter(LecturaFacturacionImpresion.validacion_manual.is_(None)).count()
    aceptados = query.filter(LecturaFacturacionImpresion.validacion_manual.is_(True)).count()
    rechazados = query.filter(LecturaFacturacionImpresion.validacion_manual.is_(False)).count()

    return {
        "periodo_anoi": periodo_anoi,
        "periodo_mes": periodo_mes,
        "total_registros": total,
        "registros_validados": validados,
        "registros_pendientes": pendientes,
        "registros_aceptados": aceptados,
        "registros_rechazados": rechazados
    }

@router.get("/exportar-excel")
def exportar_excel(
    periodo_anoi: str = Query(...),
    periodo_mes: str = Query(...),
    db: Session = Depends(get_db)
):
    registros = db.query(LecturaFacturacionImpresion).filter(
        LecturaFacturacionImpresion.periodo_anoi == periodo_anoi,
        LecturaFacturacionImpresion.periodo_mes == periodo_mes
    ).all()

    if not registros:
        raise HTTPException(
            status_code=404,
            detail=f"No hay registros para exportar en el periodo {periodo_anoi}-{periodo_mes}."
        )

    def format_bool(val):
        if val is None:
            return "PENDIENTE"
        elif val is True:
            return "VALIDO"
        elif val is False:
            return "INVALIDO"
        return val

    rows = []
    for r in registros:
        row = {
            "ID Registro": r.id_registro,
            "Año": r.periodo_anoi,
            "Mes": r.periodo_mes,
            "Tipo Registro": r.tipo_registro,
            "ID ANAM": r.id_anam,
            "VPN": r.vpn,
            "Estado": r.estado,
            "Unidad Administrativa": r.unidad_administrativa,
            "Unidad Adm II": r.unidad_administrativa_ii,
            "Serie": r.serie,
            "Modelo": r.modelo,
            
            # Campos Booleanos transformados según regla: None -> PENDIENTE, True -> VALIDO, False -> INVALIDO
            "Validación Automática": format_bool(r.validacion_automatica),
            "Validación Manual": format_bool(r.validacion_manual),
            "Fecha Válida": format_bool(r.fecha_valida),
            
            "Fecha Validación Manual": r.fecha_validacion_manual.strftime("%Y-%m-%d %H:%M:%S") if r.fecha_validacion_manual else "",
            "Observaciones Auto": r.observaciones_auto or "",
            
            # Secciones Fijas (Lecturas Iniciales)
            "Inicial Carta BN": r.lectura_inicial_carta_bn,
            "Inicial Oficio BN": r.lectura_inicial_oficio_bn,
            "Inicial Doble Carta": r.lectura_inicial_doblecarta,
            "Inicial Carta Color": r.lectura_inicial_carta_cl,
            "Inicial Oficio Color": r.lectura_inicial_oficio_cl,
            "Inicial Digitalizar": r.lectura_inicial_digitalizar,
            
            # Lecturas Finales
            "Final Carta BN": r.lectura_final_carta_bn,
            "Final Oficio BN": r.lectura_final_oficio_bn,
            "Final Doble Carta": r.lectura_final_doblecarta,
            "Final Carta Color": r.lectura_final_carta_cl,
            "Final Oficio Color": r.lectura_final_oficio_cl,
            "Final Digitalizar": r.lectura_final_digitalizar,
            
            # Volumen
            "Volumen Carta BN": r.volumen_carta_bn,
            "Volumen Oficio BN": r.volumen_oficio_bn,
            "Volumen Doble Carta": r.volumen_doblecarta,
            "Volumen Carta Color": r.volumen_carta_cl,
            "Volumen Oficio Color": r.volumen_oficio_cl,
            "Volumen Digitalizar": r.volumen_digitalizar,
            
            # Precios Unitarios
            "Precio Carta BN": float(r.precio_carta_bn or 0),
            "Precio Oficio BN": float(r.precio_oficio_bn or 0),
            "Precio Doble Carta": float(r.precio_doblecarta or 0),
            "Precio Carta Color": float(r.precio_carta_cl or 0),
            "Precio Oficio Color": float(r.precio_oficio_cl or 0),
            "Precio Digitalizar": float(r.precio_digitalizar or 0),
            
            # Importes
            "Importe Carta BN": float(r.importe_carta_bn or 0),
            "Importe Oficio BN": float(r.importe_oficio_bn or 0),
            "Importe Doble Carta": float(r.importe_doblecarta or 0),
            "Importe Carta Color": float(r.importe_carta_cl or 0),
            "Importe Oficio Color": float(r.importe_oficio_cl or 0),
            "Importe Digitalizar": float(r.importe_digitalizar or 0),
            
            # Totales
            "Subtotal": float(r.subtotal or 0),
            "IVA": float(r.iva or 0),
            "Total": float(r.total or 0),
            "Usuario Creación": r.usuario_creacion or "",
            "Fecha Creación": r.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S") if r.fecha_creacion else ""
        }
        rows.append(row)

    # Creación y Estilizado del Libro de Excel mediante openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{periodo_anoi}-{periodo_mes}"

    # Estilos de Fondo (PatternFill)
    fill_header = PatternFill(start_color="691C32", end_color="691C32", fill_type="solid")
    font_header = Font(color="FFFFFF", bold=True, size=10)

    # Reglas de Colores Solicitadas:
    # 1. Validación manual = VALIDO -> Verde claro (#E6F4EA)
    fill_verde = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")

    # 2. Validación manual = INVALIDO -> Rojo claro (#FFEBEE)
    fill_rojo = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    # 3. Validación automática = INVALIDO -> Amarillo (#FFF9C4)
    fill_amarillo = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    # 4. Validación manual = PENDIENTE y Validación automática = VALIDO -> Blanco (#FFFFFF)
    fill_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Escribir Encabezados
    headers_list = list(rows[0].keys()) if rows else []
    ws.append(headers_list)

    for col_idx in range(1, len(headers_list) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Escribir Datos y Aplicar Fondos según la Lógica de Validación
    for r_idx, (r_obj, row_data) in enumerate(zip(registros, rows), start=2):
        val_man = r_obj.validacion_manual
        val_auto = r_obj.validacion_automatica

        # Determinar el color del renglón según las prioridades definidas:
        if val_man is True:
            row_fill = fill_verde
        elif val_man is False:
            row_fill = fill_rojo
        elif val_auto is False:
            row_fill = fill_amarillo
        else:  # val_man is None and val_auto is True
            row_fill = fill_blanco

        row_values = list(row_data.values())
        ws.append(row_values)

        for col_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.fill = row_fill
            cell.font = Font(size=9)

    output = io.BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()

    fecha_ddmmyyyy = datetime.now().strftime("%d%m%Y")
    filename = f"{periodo_anoi}-{periodo_mes}-validacion-{fecha_ddmmyyyy}.xlsx"

    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Access-Control-Expose-Headers": "Content-Disposition"
    }

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@router.get("/registro-detalle")
def get_registro_detalle(
    periodo_anoi: str = Query(...),
    periodo_mes: str = Query(...),
    serie: str = Query(...),
    db: Session = Depends(get_db)
):
    registro = db.query(LecturaFacturacionImpresion).filter(
        LecturaFacturacionImpresion.periodo_anoi == periodo_anoi,
        LecturaFacturacionImpresion.periodo_mes == periodo_mes,
        LecturaFacturacionImpresion.serie == serie
    ).first()

    if not registro:
        raise HTTPException(status_code=404, detail=f"No se encontró registro para la serie '{serie}' en el periodo {periodo_anoi}-{periodo_mes}")

    columnas = db.query(LecturaFacturacionImpresionColumna).filter(
        LecturaFacturacionImpresionColumna.id_registro == registro.id_registro
    ).all()

    # Group slave columns by category (Secciones 2 a 8)
    categorias_dict: Dict[str, List[Dict[str, Any]]] = {
        "IMPRESION CARTA BN": [],
        "IMPRESION OFICIO BN": [],
        "IMPRESION DOBLECARTA BN": [],
        "DIGITALIZACION": [],
        "IMPRESION CARTA COLOR": [],
        "IMPRESION OFICIO COLOR": [],
        "IMPRESION DOBLECARTA COLOR": []
    }

    for col in columnas:
        cat = col.categoria
        if cat not in categorias_dict:
            categorias_dict[cat] = []
        categorias_dict[cat].append({
            "id_columna": col.id_columna,
            "nombre_columna": col.nombre_columna,
            "valor_columna": col.valor_columna,
            "es_valido": col.es_valido
        })

    return {
        "id_registro": registro.id_registro,
        "periodo_anoi": registro.periodo_anoi,
        "periodo_mes": registro.periodo_mes,
        "tipo_registro": registro.tipo_registro,

        # Equipo
        "id_anam": registro.id_anam,
        "vpn": registro.vpn,
        "estado": registro.estado,
        "unidad_administrativa": registro.unidad_administrativa,
        "unidad_administrativa_ii": registro.unidad_administrativa_ii,
        "serie": registro.serie,
        "modelo": registro.modelo,

        # Estatus de Validación (Pendiente = None, Correcto = True, Rechazado = False)
        "validacion_automatica": registro.validacion_automatica,
        "validacion_manual": registro.validacion_manual,
        "fecha_validacion_manual": registro.fecha_validacion_manual.isoformat() if registro.fecha_validacion_manual else None,
        "observaciones_auto": registro.observaciones_auto,
        "fecha_valida": registro.fecha_valida,
        "usuario_validacion": registro.usuario_creacion,

        # Categorías Variables (Secciones 2 a 8)
        "categorias_variables": categorias_dict,

        # Secciones Fijas (9 a 14)
        "secciones_fijas": {
            "lecturas_iniciales": {
                "carta_bn": registro.lectura_inicial_carta_bn,
                "oficio_bn": registro.lectura_inicial_oficio_bn,
                "doblecarta": registro.lectura_inicial_doblecarta,
                "carta_cl": registro.lectura_inicial_carta_cl,
                "oficio_cl": registro.lectura_inicial_oficio_cl,
                "digitalizar": registro.lectura_inicial_digitalizar
            },
            "lecturas_finales": {
                "carta_bn": registro.lectura_final_carta_bn,
                "oficio_bn": registro.lectura_final_oficio_bn,
                "doblecarta": registro.lectura_final_doblecarta,
                "carta_cl": registro.lectura_final_carta_cl,
                "oficio_cl": registro.lectura_final_oficio_cl,
                "digitalizar": registro.lectura_final_digitalizar
            },
            "volumen": {
                "carta_bn": registro.volumen_carta_bn,
                "oficio_bn": registro.volumen_oficio_bn,
                "doblecarta": registro.volumen_doblecarta,
                "carta_cl": registro.volumen_carta_cl,
                "oficio_cl": registro.volumen_oficio_cl,
                "digitalizar": registro.volumen_digitalizar
            },
            "precios": {
                "carta_bn": float(registro.precio_carta_bn or 0),
                "oficio_bn": float(registro.precio_oficio_bn or 0),
                "doblecarta": float(registro.precio_doblecarta or 0),
                "carta_cl": float(registro.precio_carta_cl or 0),
                "oficio_cl": float(registro.precio_oficio_cl or 0),
                "digitalizar": float(registro.precio_digitalizar or 0)
            },
            "importes": {
                "carta_bn": float(registro.importe_carta_bn or 0),
                "oficio_bn": float(registro.importe_oficio_bn or 0),
                "doblecarta": float(registro.importe_doblecarta or 0),
                "carta_cl": float(registro.importe_carta_cl or 0),
                "oficio_cl": float(registro.importe_oficio_cl or 0),
                "digitalizar": float(registro.importe_digitalizar or 0)
            },
            "totales": {
                "subtotal": float(registro.subtotal or 0),
                "iva": float(registro.iva or 0),
                "total": float(registro.total or 0)
            }
        }
    }

@router.post("/validar-manual")
def validar_manual(req: ValidacionManualRequest, db: Session = Depends(get_db)):
    registro = db.query(LecturaFacturacionImpresion).filter(
        LecturaFacturacionImpresion.id_registro == req.id_registro
    ).first()

    if not registro:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    # Actualizar estado de columnas esclavas
    for col_dto in req.columnas_estados:
        db.query(LecturaFacturacionImpresionColumna).filter(
            LecturaFacturacionImpresionColumna.id_columna == col_dto.id_columna
        ).update({"es_valido": col_dto.es_valido})

    # Actualizar maestro con validación manual y fecha
    registro.fecha_valida = req.fecha_valida
    registro.validacion_manual = (req.decision.upper() == "ACEPTAR")
    registro.fecha_validacion_manual = datetime.now()

    db.commit()
    return {"status": "success", "mensaje": "Validación manual guardada correctamente"}
