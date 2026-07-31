import io
import re
import openpyxl
from typing import Dict, List, Tuple, Any, Optional

def extract_period_from_filename(filename: str) -> Tuple[Optional[str], str]:
    """
    Extract (periodo_mes, periodo_anoi) from filename.
    Examples:
      '01-abril-2026.xlsx' -> ('ABR', '2026')
      '00-iniciales.xlsx'  -> ('INI', '2026')
    """
    fn = filename.lower()
    months_map = {
        "enero": "ENE", "ene": "ENE",
        "febrero": "FEB", "feb": "FEB",
        "marzo": "MAR", "mar": "MAR",
        "abril": "ABR", "abr": "ABR",
        "mayo": "MAY", "may": "MAY",
        "junio": "JUN", "jun": "JUN",
        "julio": "JUL", "jul": "JUL",
        "agosto": "AGO", "ago": "AGO",
        "septiembre": "SEP", "sep": "SEP",
        "octubre": "OCT", "oct": "OCT",
        "noviembre": "NOV", "nov": "NOV",
        "diciembre": "DIC", "dic": "DIC",
        "inicial": "INI", "iniciales": "INI", "ini": "INI"
    }

    detected_mes = None
    for key, code in months_map.items():
        if key in fn:
            detected_mes = code
            break

    years = re.findall(r"20\d{2}", fn)
    detected_anoi = years[0] if years else "2026"

    return detected_mes, detected_anoi


def parse_excel_file(file_contents: bytes, filename: str) -> Tuple[int, List[Dict[str, Any]]]:
    """
    Parse an Excel report file (.xlsx) and extract:
    1. Row 1 & 2 merged header categories.
    2. Column header mapping across all 14 sections.
    3. Row data records with master section fields and slave variable column records.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_contents), data_only=True)
    ws = wb.active

    max_c = ws.max_column
    max_r = ws.max_row

    # Unmerge / fill header cells for rows 1 and 2
    row1 = [ws.cell(1, c).value for c in range(1, max_c + 1)]
    row2 = [ws.cell(2, c).value for c in range(1, max_c + 1)]

    for rng in ws.merged_cells.ranges:
        v1 = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, min(rng.max_row + 1, 3)):
            for c in range(rng.min_col, rng.max_col + 1):
                if r == 1:
                    row1[c - 1] = v1
                elif r == 2:
                    row2[c - 1] = v1

    # Detect header row (row 2 in monthly files or row 3/4)
    header_row_idx = 2
    if ws.cell(4, 1).value == "ID-ANAM":
        header_row_idx = 4
    elif ws.cell(3, 1).value == "ID-ANAM":
        header_row_idx = 3

    headers = [ws.cell(header_row_idx, c).value for c in range(1, max_c + 1)]

    # Validate essential metadata column headers
    raw_headers_str = " ".join([str(h).upper() for h in headers if h])
    if "SERIE" not in raw_headers_str or "MODELO" not in raw_headers_str or "ID-ANAM" not in raw_headers_str:
        raise ValueError("El formato de los primeros renglones no coincide con el formato esperado.")

    is_inicial_file = "00-inicial" in filename.lower() or "inicial" in filename.lower()

    # Classify columns into sections
    column_maps = []
    for c in range(max_c):
        r1_val = str(row1[c]).strip() if row1[c] else ""
        r2_val = str(row2[c]).strip() if row2[c] else ""
        h_val = str(headers[c]).strip() if headers[c] else f"COL_{c+1}"

        sec_name, sub_name = _classify_column(r1_val, r2_val, h_val, c + 1, is_inicial_file)
        column_maps.append({
            "col_idx": c + 1,
            "section": sec_name,
            "sub_name": sub_name,
            "raw_header": h_val,
            "r1": r1_val,
            "r2": r2_val
        })

    # Read data records starting from row after header_row_idx
    records = []
    for r in range(header_row_idx + 1, max_r + 1):
        row_cells = [ws.cell(r, c).value for c in range(1, max_c + 1)]
        if not any(row_cells):
            continue

        record = {
            "equipo": {},
            "lecturas_iniciales": {},
            "lecturas_finales": {},
            "volumen": {},
            "precios": {},
            "importes": {},
            "totales": {},
            "columnas_variables": []  # List of dicts for slave table
        }

        # Check Serie first (required)
        serie_val = None
        for col_info in column_maps:
            if col_info["section"] == "EQUIPO" and col_info["sub_name"].upper() == "SERIE":
                val = ws.cell(r, col_info["col_idx"]).value
                if val is not None and str(val).strip() != "" and not str(val).startswith("#"):
                    serie_val = str(val).strip()
                break

        if not serie_val:
            continue

        for col_info in column_maps:
            c_idx = col_info["col_idx"]
            val = ws.cell(r, c_idx).value
            sec = col_info["section"]
            sub = col_info["sub_name"]

            if sec == "EQUIPO":
                record["equipo"][sub.upper()] = val

            elif sec == "LECTURAS INICIALES":
                record["lecturas_iniciales"][sub] = _to_int(val)

            elif sec == "LECTURAS FINALES":
                record["lecturas_finales"][sub] = _to_int(val)

            elif sec == "VOLUMEN":
                record["volumen"][sub] = _to_int(val)

            elif sec == "PRECIOS":
                record["precios"][sub] = _to_float(val)

            elif sec == "IMPORTES FACTURACION":
                record["importes"][sub] = _to_float(val)

            elif sec == "TOTALES":
                record["totales"][sub] = _to_float(val)

            elif sec.startswith("IMPRESION") or sec == "DIGITALIZACION":
                # Variable column section (2 to 8)
                if val is not None and str(val).strip() != "" and not str(val).startswith("#"):
                    record["columnas_variables"].append({
                        "serie": serie_val,
                        "categoria": sec,
                        "nombre_columna": col_info["raw_header"],
                        "valor_columna": str(val).strip(),
                        "valor_num": _to_int(val),
                        "es_valido": None
                    })

        records.append(record)

    return len(records), records


def _classify_column(r1_val: str, r2_val: str, h_val: str, col_idx: int, is_inicial_file: bool) -> Tuple[str, str]:
    r1 = r1_val.upper()
    r2 = r2_val.upper()
    h = h_val.upper()

    # 1. EQUIPO
    if h in ["ID-ANAM", "VPN", "ESTADO", "UNIDAD ADMINISTRATIVAS", "UNIDAD ADMINISTRATIVAS II", "SERIE", "MODELO"]:
        return "EQUIPO", h

    # Helper function for sub-category matching
    def _match_sub(h_text: str) -> str:
        if "DOBLECARTA" in h_text or "DOBLE CARTA" in h_text:
            return "DOBLECARTA"
        elif "DIGITALIZ" in h_text:
            return "DIGITALIZACION"
        elif "OFICIO" in h_text and ("BN" in h_text or "BLACK" in h_text):
            return "OFICIO BN"
        elif "OFICIO" in h_text and "COLOR" in h_text:
            return "OFICIO COLOR"
        elif "CARTA" in h_text and ("BN" in h_text or "BLACK" in h_text):
            return "CARTA BN"
        elif "CARTA" in h_text and "COLOR" in h_text:
            return "CARTA COLOR"
        return "DIGITALIZACION"

    # 9. LECTURAS INICIALES in 00-iniciales.xlsx (cols 128-133)
    if is_inicial_file and 128 <= col_idx <= 133:
        return "LECTURAS INICIALES", _match_sub(h)

    # Fixed sections in monthly billing files
    for sec_name in ["LECTURAS INICIALES", "LECTURAS FINALES", "VOLUMEN", "PRECIOS", "IMPORTES FACTURACION"]:
        if sec_name in r2 or sec_name in r1:
            return sec_name, _match_sub(h)

    # 14. TOTALES
    if "SUBTOTAL" in h:
        return "TOTALES", "SUBTOTAL"
    if "IVA" in h:
        return "TOTALES", "IVA"
    if "TOTAL" in h and "SUBTOTAL" not in h:
        return "TOTALES", "TOTAL"

    # Variable Sections (2 to 8) - Check DOBLECARTA before CARTA to avoid false prefix match!
    cat_str = (r1 + " " + r2).strip()
    if "DOBLECARTA BN" in cat_str or "DOBLE CARTA BN" in cat_str:
        return "IMPRESION DOBLECARTA BN", h_val
    if "DOBLECARTA CL" in cat_str or "DOBLECARTA COLOR" in cat_str or "DOBLE CARTA COLOR" in cat_str:
        return "IMPRESION DOBLECARTA COLOR", h_val
    if "CARTA BN" in cat_str:
        return "IMPRESION CARTA BN", h_val
    if "OFICIO BN" in cat_str:
        return "IMPRESION OFICIO BN", h_val
    if "ENVIAR" in cat_str or "DIGITALIZ" in cat_str:
        return "DIGITALIZACION", h_val
    if "CARTA CL" in cat_str or "CARTA COLOR" in cat_str:
        return "IMPRESION CARTA COLOR", h_val
    if "OFICIO CL" in cat_str or "OFICIO COLOR" in cat_str:
        return "IMPRESION OFICIO COLOR", h_val

    return "UNKNOWN", h_val


def _to_int(val: Any) -> int:
    if val is None or val == "" or str(val).startswith("#"):
        return 0
    try:
        return int(round(float(val)))
    except (ValueError, TypeError):
        return 0


def _to_float(val: Any) -> float:
    if val is None or val == "" or str(str(val)).startswith("#"):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0
